from __future__ import annotations

import argparse
import json
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data"
DEFAULT_INPUT = DATA_DIR / "master.xlsx"
EXCLUDED_HEADERS = {
    "略称ＣＤ",
    "分類ＣＤ",
    "分類名",
    "倉庫№",
    "倉庫名",
    "取引先CD",
    "取引先名",
    "直送フラグ",
    "勘定科目",
    "勘定科目名",
    "内訳科目",
    "内訳科目名",
    "発注点",
    "発注量",
    "最低発注数",
    "所属単位",
    "所属単位名",
    "単価使用区分",
    "単価使用区分名",
    "単価表示区分",
    "単位CD",
    "活動区分",
}


def convert_cell_value(value: Any) -> Any:
    """Convert openpyxl cell value to JSON-serializable value."""
    if value is None:
        return None
    if isinstance(value, (datetime, date, time)):
        # Use ISO-8601 formats
        if isinstance(value, time):
            return value.strftime("%H:%M:%S")
        return value.isoformat()
    if isinstance(value, float):
        # Convert floats that are effectively integers to int for cleaner JSON
        if value.is_integer():
            return int(value)
        return float(value)
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, bool):
        return bool(value)
    # Fallback: best-effort string conversion
    return str(value)


def normalize_headers(headers: Iterable[Any]) -> List[str]:
    normalized: List[str] = []
    for index, h in enumerate(headers, start=1):
        name = ""
        if h is not None:
            name = str(h).strip()
        if not name:
            # Skip empty headers (no key)
            normalized.append("")
        else:
            normalized.append(name)
    return normalized


def sheet_to_records(ws) -> List[Dict[str, Any]]:
    """Convert a worksheet to list of dict records using row 1 as headers."""
    header_rows = ws.iter_rows(min_row=1, max_row=1, values_only=True)
    try:
        raw_headers = next(header_rows)
    except StopIteration:
        return []

    headers = normalize_headers(raw_headers or [])
    if not any(h for h in headers):
        # No usable headers
        return []

    records: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        # Skip completely empty rows
        if all(cell is None or (isinstance(cell, str) and cell.strip() == "") for cell in row):
            continue

        record: Dict[str, Any] = {}
        for idx, header in enumerate(headers):
            if not header or header in EXCLUDED_HEADERS:
                # Ignore columns with empty header
                continue
            value = row[idx] if idx < len(row) else None
            record[header] = convert_cell_value(value)
        # Skip rows that produced no fields (e.g., all headers empty)
        if record:
            records.append(record)
    return records


def convert_workbook_to_json(
    input_path: Path,
    output_dir: Path,
    target_sheets: Optional[List[str]] = None,
) -> Dict[str, Path]:
    """
    Convert each worksheet to a JSON file named '<sheet>.json' in output_dir.
    Returns a mapping of sheet name to output file path.
    """
    if not input_path.exists():
        raise FileNotFoundError(f"Excelファイルが見つかりません: {input_path}")
    output_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(filename=input_path, data_only=True, read_only=True)

    written: Dict[str, Path] = {}
    for sheet_name in wb.sheetnames:
        if target_sheets and sheet_name not in target_sheets:
            continue
        ws = wb[sheet_name]
        records = sheet_to_records(ws)
        out_path = output_dir / f"{sheet_name}.json"
        with out_path.open("w", encoding="utf-8") as f:
            json.dump(records, f, ensure_ascii=False, indent=2)
        written[sheet_name] = out_path
    return written


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="master.xlsx を JSON に変換します（各シート→個別JSON）。")
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT,
        help="入力のExcelファイルパス（既定: data/master.xlsx）",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DATA_DIR,
        help="JSON出力ディレクトリ（既定: data/）",
    )
    parser.add_argument(
        "--sheet",
        action="append",
        default=None,
        help="変換対象シート名（複数指定可）。未指定なら全シートを変換。",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    written = convert_workbook_to_json(args.input, args.output_dir, args.sheet)
    if not written:
        print("変換対象がありませんでした。見出し行やシート指定をご確認ください。")
        return
    print("変換完了:")
    for name, path in written.items():
        print(f"- {name}: {path}")


if __name__ == "__main__":
    main()

